import re
from collections import defaultdict
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import sys
import os

def parse_report(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()

    monthly_data = defaultdict(lambda: defaultdict(int))
    total_pieces = 0
    total_shifts = 0

    current_month = None
    looking_for_date = True
    line_number = None
    dates = []

    i = 0
    while i < len(lines):
        line = lines[i]
        
        if looking_for_date:
            # Match lines with the date format m/d/yy or m/d/yyyy, allowing for leading whitespace
            date_match = re.match(r'\s*(\d{1,2}/\d{1,2}/\d{2,4})', line)
            if date_match:
                date_str = date_match.group(1)
                try:
                    date = datetime.strptime(date_str, '%m/%d/%y').date()
                except ValueError:
                    date = datetime.strptime(date_str, '%m/%d/%Y').date()
                current_month = date.strftime('%Y-%m')
                dates.append(date)
                looking_for_date = False  # Stop looking for date and start looking for "Total Machine # Shift #"
        else:
            # Match lines with "Total Machine # Shift #"
            machine_shift_match = re.search(r'Total Machine (\d+)  Shift \d+', line)
            if machine_shift_match:
                line_number = machine_shift_match.group(1)
                total_shifts += 1
                # Look for "Pcs:" in the next line
                if i + 1 < len(lines):
                    next_line = lines[i + 1]
                    pcs_match = re.search(r'Pcs:\s+(\d+)', next_line)
                    if pcs_match:
                        pieces = int(pcs_match.group(1))
                        total_pieces += pieces
                        monthly_data[current_month]['pieces'] += pieces
                        monthly_data[current_month]['shifts'] += 1
                i += 1  # Skip the next line since we've already processed it
                looking_for_date = True  # Start looking for date again
        i += 1

    date_range = f"{min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}" if dates else "No Date Range"
    return monthly_data, total_pieces, total_shifts, line_number, date_range

def generate_report(monthly_data, total_pieces, total_shifts):
    report = []
    total_hours = total_shifts * 7.25

    report.append(f"Total Pieces: {total_pieces}")
    report.append(f"Total Shifts: {total_shifts}")
    if total_hours > 0:
        report.append(f"Total Pieces per Hour: {total_pieces / total_hours:.2f}")
    else:
        report.append("Total Pieces per Hour: N/A (No shifts detected)")

    report.append("\nMonthly Data:")
    for month, info in sorted(monthly_data.items()):
        monthly_hours = info['shifts'] * 7.25
        pieces_per_hour = info['pieces'] / monthly_hours if monthly_hours > 0 else 0
        report.append(f"Month: {month}, Pieces: {info['pieces']}, Shifts: {info['shifts']}, Pieces per Hour: {pieces_per_hour:.2f}")

    return "\n".join(report)

def generate_excel_report(monthly_data, total_pieces, total_shifts, line_number, date_range, file_path):
    total_hours = total_shifts * 7.25
    overall_data = {
        'Total Pieces': [total_pieces],
        'Total Shifts': [total_shifts],
        'Total Pieces per Hour': [total_pieces / total_hours if total_hours > 0 else 'N/A']
    }
    overall_df = pd.DataFrame(overall_data)

    monthly_report = []
    for month, info in sorted(monthly_data.items()):
        monthly_hours = info['shifts'] * 7.25
        pieces_per_hour = info['pieces'] / monthly_hours if monthly_hours > 0 else 0
        monthly_report.append({
            'Month': month,
            'Pieces': info['pieces'],
            'Shifts': info['shifts'],
            'Pieces per Hour': pieces_per_hour
        })
    monthly_df = pd.DataFrame(monthly_report)

    with pd.ExcelWriter(file_path) as writer:
        overall_df.to_excel(writer, sheet_name='Overall Data', index=False)
        monthly_df.to_excel(writer, sheet_name='Monthly Data', index=False)

    # Load the workbook and add Line # and Date Range to the first sheet
    workbook = load_workbook(file_path)
    worksheet_overall = workbook['Overall Data']
    worksheet_overall['A1'] = f"Line: {line_number}"
    worksheet_overall['A2'] = f"Date Range: {date_range}"

    # Apply formatting to Overall Data sheet
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    for col in worksheet_overall.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet_overall.column_dimensions[column].width = adjusted_width

    for cell in worksheet_overall["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Apply formatting to Monthly Data sheet
    worksheet_monthly = workbook['Monthly Data']

    for col in worksheet_monthly.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet_monthly.column_dimensions[column].width = adjusted_width

    for cell in worksheet_monthly["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    workbook.save(file_path)

if __name__ == "__main__":
    import os
    import sys

    # Check if a file path is provided as a command-line argument
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # Prompt the user for the file path
        file_path = input("Please enter the path to the text file: ").strip()
    
    if not os.path.isfile(file_path):
        print(f"File not found: {file_path}")
    else:
        monthly_data, total_pieces, total_shifts, line_number, date_range = parse_report(file_path)
        report = generate_report(monthly_data, total_pieces, total_shifts)
        print(report)

        excel_file_path = f'Line_{line_number}_{date_range.replace(" ", "_").replace(":", "-")}.xlsx'  # Replace with your desired Excel file path
        generate_excel_report(monthly_data, total_pieces, total_shifts, line_number, date_range, excel_file_path)
        print(f"Excel report generated: {excel_file_path}")